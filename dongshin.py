"""
AI 대학 합격 가능성 분석기 (범용 엑셀 자동 파싱 버전)

특정 대학 양식에 종속되지 않고,
- 제목이 여러 줄인 엑셀
- 병합 셀이 많은 엑셀
- 헤더가 여러 줄(Multi Header)인 엑셀
- 컬럼명이 대학마다 제각각인 엑셀
을 최대한 자동으로 인식해서 다음 5개 표준 컬럼으로 정리한다.

    학과, 평균등급, 최종합격등급, 경쟁률, 충원율

주의: 완전히 자유 양식인 엑셀을 100% 자동 인식하는 것은 불가능하므로,
      아래 로직은 "대부분의 입시결과 엑셀에서 통하는 휴리스틱(경험적 규칙)"이다.
      실패 시 사용자가 원인을 알 수 있도록 각 단계에서 예외 메시지를 명확히 남긴다.
"""

import io
import re

import openpyxl
import pandas as pd
import streamlit as st

# =========================================================================
# 0. 설정값
# =========================================================================

# 표준 컬럼명과, 실제 엑셀에서 그 컬럼을 가리킬 수 있는 여러 표현들.
# 필요하면 이 목록에 대학별로 쓰는 다른 표현을 계속 추가하면 된다.
# (앞쪽에 있는 키워드일수록 우선적으로 매칭을 시도한다)
COLUMN_KEYWORDS = {
    "학과": ["모집단위명", "모집단위", "학과명", "학과"],
    "평균등급": ["학생부평균등급", "평균등급", "평균내신", "내신평균", "평균"],
    # "70%cut"처럼 영문 cut을 쓰는 대학도 있고 "70컷"처럼 한글로 쓰는 대학도 있어 둘 다 포함
    "최종합격등급": ["최종합격등급", "70cut", "70컷", "최종컷", "합격컷", "컷"],
    "경쟁률": ["경쟁률"],
    # "충원"만 넣으면 "차수별 충원 합격 예비순위 1차"처럼 전혀 다른 컬럼(회차별 인원)에도
    # 잘못 매칭될 수 있어, 실제로 "충원율"이라는 표현이 있는 경우만 매칭한다.
    # 명시적인 충원율 컬럼이 없으면 아래 compute_supplementary_rate()에서 추정치를 계산한다.
    "충원율": ["충원율", "추가합격율"],
}

# 충원율 컬럼이 없을 때, "모집인원"과 "충원 최종 예비순위(인원)"으로
# 충원율(%)을 역산하기 위한 보조 컬럼 키워드.
# "전형구분명"은 같은 학과라도 학생부교과/학생부종합/논술 등 전형에 따라
# 합격선이 크게 다르므로, 학과 중복을 구분하는 용도로 함께 찾아둔다.
AUX_COLUMN_KEYWORDS = {
    "모집인원": ["모집인원"],
    "충원최종인원": ["충원합격최종예비순위", "최종예비순위", "충원최종예비순위", "최종충원순위"],
    "전형구분": ["전형구분명", "전형명", "전형유형", "전형구분"],
}

REQUIRED_CANONICAL_COLS = ["학과", "평균등급", "최종합격등급"]   # 없으면 분석 불가
NUMERIC_CANONICAL_COLS = ["평균등급", "최종합격등급", "경쟁률", "충원율"]


# =========================================================================
# 1. 엑셀 원본 읽기 + 병합 셀 해제
# =========================================================================
def load_raw_rows(uploaded_file):
    """
    업로드된 엑셀 파일(첫 번째 시트)을 openpyxl로 읽는다.
    병합 셀은 모두 해제한 뒤, 좌상단(원래 값이 있던 셀)의 값으로
    병합 범위 전체를 채워서 pandas 없이도 2차원 리스트로 다룰 수 있게 한다.
    (병합 셀 문제 처리 - 요구사항 3)
    """
    try:
        file_bytes = uploaded_file.getvalue()
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]  # 첫 번째 시트만 사용
    except Exception as e:
        raise ValueError(f"엑셀 파일을 여는 중 오류가 발생했습니다: {e}")

    # 병합 셀 목록을 미리 리스트로 복사해둔다 (반복 중 해제하면서 목록이 바뀌기 때문)
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row = merged_range.min_col, merged_range.min_row
        max_col, max_row = merged_range.max_col, merged_range.max_row
        top_left_value = worksheet.cell(row=min_row, column=min_col).value

        worksheet.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                worksheet.cell(row=r, column=c).value = top_left_value

    raw_rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]

    if not raw_rows:
        raise ValueError("엑셀 시트에 데이터가 없습니다.")

    return raw_rows


# =========================================================================
# 2. 실제 데이터가 시작되는 행 찾기 (요구사항 1)
# =========================================================================
def _is_numeric_like(value):
    """셀 값이 숫자로 볼 수 있는지 확인 ('%', ',' 등이 섞여 있어도 인정)."""
    if value is None:
        return False
    try:
        float(str(value).replace("%", "").replace(",", "").strip())
        return True
    except ValueError:
        return False


def find_data_start_row(raw_rows, max_search_rows=40, numeric_ratio_threshold=0.3):
    """
    위에서부터 각 행을 검사하며, '실제 데이터 행'이 시작되는 위치를 찾는다.

    판단 기준:
    - 채워진(빈 값이 아닌) 셀이 2개 이상이고
    - 그 중 숫자로 변환 가능한 셀의 비율이 threshold 이상이면
    -> 데이터 행으로 간주한다. (제목/헤더 행은 대부분 텍스트라 비율이 낮음)
    """
    search_range = raw_rows[:max_search_rows]

    for idx, row in enumerate(search_range):
        non_null = [v for v in row if v is not None and str(v).strip() != ""]
        if len(non_null) < 2:
            continue

        numeric_count = sum(1 for v in non_null if _is_numeric_like(v))
        ratio = numeric_count / len(non_null)

        if ratio >= numeric_ratio_threshold:
            return idx

    raise ValueError(
        "데이터가 시작되는 행을 찾지 못했습니다. "
        "엑셀 형식이 예상과 많이 다를 수 있습니다."
    )


# =========================================================================
# 3. 헤더 행 추출 + Multi Header 합치기 (요구사항 2)
# =========================================================================
def extract_header_rows(raw_rows, data_start_idx, max_header_rows=4):
    """
    data_start_idx 바로 위에서부터 위로 올라가며 헤더로 보이는 행들을 모은다.
    - 채워진 셀이 2개 이상인 행 -> 헤더 행으로 인정
    - 채워진 셀이 1개 이하인 행(보통 제목행: "2025학년도 수시 결과" 등) -> 헤더가 아니므로 제외하고 탐색 중단
    반환값은 위->아래 순서로 정렬된 리스트.
    """
    header_rows = []
    idx = data_start_idx - 1
    collected = 0

    while idx >= 0 and collected < max_header_rows:
        row = raw_rows[idx]
        non_null = [v for v in row if v is not None and str(v).strip() != ""]

        if len(non_null) >= 2:
            header_rows.insert(0, row)  # 위->아래 순서 유지를 위해 맨 앞에 삽입
            collected += 1
        else:
            # 헤더를 이미 하나라도 찾은 상태에서 텅 빈(제목) 행을 만나면 그 위는 보지 않는다
            if header_rows:
                break
        idx -= 1

    if not header_rows:
        raise ValueError("헤더 행을 찾지 못했습니다. 엑셀 형식을 확인해주세요.")

    return header_rows


def normalize_text(text):
    """
    비교를 쉽게 하기 위해 공백/줄바꿈/기호를 제거하고 소문자로 통일한다.
    예: "학생부 평균등급\n(전체)" -> "학생부평균등급전체"
    """
    if text is None:
        return ""
    text = str(text).replace("\n", "").replace("\r", "")
    text = re.sub(r"[^0-9A-Za-z가-힣]", "", text)
    return text.lower()


def flatten_multi_header(header_rows):
    """
    여러 줄로 나뉜 헤더(Multi Header)를 각 열마다 한 줄로 합친다.
    같은 열에서 위->아래로 값을 이어붙이되, 바로 위 칸과 값이 같으면
    (병합으로 인해 값이 반복되는 경우) 한 번만 사용한다.

    예) 1행: "학생부" | 2행: "평균등급"  ->  "학생부평균등급"  (요구사항 2)
    """
    num_cols = max(len(row) for row in header_rows)
    flattened_names = []

    for col_idx in range(num_cols):
        parts = []
        prev_value = None
        for row in header_rows:
            value = row[col_idx] if col_idx < len(row) else None
            if value is None or str(value).strip() == "":
                continue
            value = str(value).strip()
            if value != prev_value:
                parts.append(value)
            prev_value = value
        flattened_names.append("".join(parts))

    return flattened_names


# =========================================================================
# 4. 표준 컬럼과 자동 매칭 (요구사항 4)
# =========================================================================
def match_columns(column_names, keyword_map=None, used_idx=None):
    """
    실제 컬럼명 목록에서 표준 컬럼(기본값: 학과/평균등급/최종합격등급/경쟁률/충원율)에
    해당하는 열의 인덱스를 찾아 딕셔너리로 반환한다.
    이미 다른 표준 컬럼(또는 이전 매칭 단계)에 배정된 열은 중복으로 배정하지 않도록
    used_idx를 넘겨받아 이어서 사용할 수 있다. (보조 컬럼 매칭 시 재사용)

    매칭되지 않은 표준 컬럼은 값이 None으로 남는다.
    반환값: (mapping, used_idx) - used_idx는 이후 다른 매칭 단계에 그대로 넘길 수 있다.
    """
    if keyword_map is None:
        keyword_map = COLUMN_KEYWORDS
    used_idx = set() if used_idx is None else set(used_idx)

    normalized = [normalize_text(name) for name in column_names]
    mapping = {}

    for canonical, keywords in keyword_map.items():
        found_idx = None
        for keyword in keywords:
            norm_keyword = normalize_text(keyword)
            if not norm_keyword:
                continue
            for idx, col_norm in enumerate(normalized):
                if idx in used_idx:
                    continue
                if norm_keyword in col_norm:
                    found_idx = idx
                    break
            if found_idx is not None:
                break
        mapping[canonical] = found_idx
        if found_idx is not None:
            used_idx.add(found_idx)

    return mapping, used_idx


# =========================================================================
# 5. 표준 DataFrame 생성 (요구사항 5)
# =========================================================================
def _to_number(value):
    """'%', ',' 등이 섞인 값도 최대한 숫자로 변환한다. 실패하면 None(결측치) 처리."""
    if value is None:
        return None
    try:
        cleaned = str(value).replace("%", "").replace(",", "").strip()
        if cleaned == "":
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def build_clean_dataframe(raw_rows, data_start_idx, column_mapping, aux_mapping=None):
    """
    매칭된 컬럼 인덱스를 이용해 실제 데이터 영역에서
    [학과, 평균등급, 최종합격등급, 경쟁률, 충원율] 5개 컬럼으로
    이루어진 표준 DataFrame을 만든다.

    충원율 컬럼을 직접 찾지 못한 경우, aux_mapping에 "모집인원"과
    "충원최종인원"이 모두 있으면 충원율(%) = 충원최종인원 / 모집인원 * 100 으로
    추정치를 계산해 채워 넣는다. (요구사항의 "충원율" 항목을 최대한 채우기 위함)
    """
    missing = [c for c in REQUIRED_CANONICAL_COLS if column_mapping.get(c) is None]
    if missing:
        raise ValueError(
            f"다음 필수 항목에 해당하는 컬럼을 엑셀에서 찾지 못했습니다: {missing}. "
            "컬럼명이 특이한 경우 COLUMN_KEYWORDS 목록에 표현을 추가해야 합니다."
        )

    data_rows = raw_rows[data_start_idx:]
    columns_data = {}

    for canonical in ["학과", "평균등급", "최종합격등급", "경쟁률", "충원율"]:
        idx = column_mapping.get(canonical)
        if idx is None:
            # 경쟁률/충원율처럼 선택 항목은 없어도 빈 값으로 채워서 진행
            columns_data[canonical] = [None] * len(data_rows)
        else:
            columns_data[canonical] = [
                row[idx] if idx < len(row) else None for row in data_rows
            ]

    # 전형구분(학생부교과/학생부종합/논술 등)이 있으면 함께 담아둔다.
    # 같은 학과라도 전형에 따라 합격선이 크게 다르므로, 학과 중복을 구분하는 데 쓰인다.
    has_admission_type = False
    if aux_mapping:
        type_idx = aux_mapping.get("전형구분")
        if type_idx is not None:
            columns_data["전형구분"] = [
                row[type_idx] if type_idx < len(row) else None for row in data_rows
            ]
            has_admission_type = True

    # 충원율 컬럼을 직접 찾지 못했다면, 보조 컬럼(모집인원/충원최종인원)으로 추정치를 계산해둔다.
    # data_rows와 같은 길이로 만들어 두면, 아래에서 학과 빈 행을 걸러낼 때 동일한 마스크를 그대로 적용할 수 있다.
    used_derived_rate = False
    if aux_mapping and column_mapping.get("충원율") is None:
        recruit_idx = aux_mapping.get("모집인원")
        final_idx = aux_mapping.get("충원최종인원")
        if recruit_idx is not None and final_idx is not None:
            derived_rate = []
            for row in data_rows:
                recruit_val = _to_number(row[recruit_idx] if recruit_idx < len(row) else None)
                final_val = _to_number(row[final_idx] if final_idx < len(row) else None)
                if recruit_val and final_val is not None and recruit_val != 0:
                    derived_rate.append(round(final_val / recruit_val * 100, 1))
                else:
                    derived_rate.append(None)
            columns_data["충원율"] = derived_rate
            used_derived_rate = True

    df = pd.DataFrame(columns_data)

    # 학과명이 비어있는 행(하단 각주, 합계, 안내문 등)은 제거 (다른 컬럼도 같은 마스크로 함께 걸러짐)
    df["학과"] = df["학과"].apply(lambda v: "" if v is None else str(v).strip())
    df = df[(df["학과"] != "") & (df["학과"].str.lower() != "none")]

    # 숫자 컬럼 변환 (충원율은 이미 숫자로 계산된 경우도 있으므로 그대로 통과)
    for col in NUMERIC_CANONICAL_COLS:
        df[col] = df[col].apply(_to_number)

    if has_admission_type:
        df["전형구분"] = df["전형구분"].apply(lambda v: "" if v is None else str(v).strip())

    df = df.reset_index(drop=True)

    if df.empty:
        raise ValueError("데이터 영역에서 유효한 행을 찾지 못했습니다.")

    df.attrs["충원율_추정치_사용"] = used_derived_rate
    df.attrs["전형구분_존재"] = has_admission_type
    return df


# =========================================================================
# 6. 전체 파싱 파이프라인
# =========================================================================
def parse_admission_excel(uploaded_file):
    """
    업로드된 엑셀 파일을 자동으로 분석하여 표준화된
    입시결과 DataFrame과, 디버깅용 정보(감지된 컬럼명/매칭 결과)를 함께 반환한다.
    실패하면 ValueError를 발생시키며 원인을 메시지에 담는다.
    """
    raw_rows = load_raw_rows(uploaded_file)
    data_start_idx = find_data_start_row(raw_rows)
    header_rows = extract_header_rows(raw_rows, data_start_idx)
    column_names = flatten_multi_header(header_rows)

    column_mapping, used_idx = match_columns(column_names, COLUMN_KEYWORDS)
    aux_mapping, _ = match_columns(column_names, AUX_COLUMN_KEYWORDS, used_idx)

    df = build_clean_dataframe(raw_rows, data_start_idx, column_mapping, aux_mapping)

    debug_info = {
        "data_start_row_in_excel": data_start_idx + 1,  # 1-based로 사용자에게 안내
        "detected_column_names": column_names,
        "column_mapping": column_mapping,
        "aux_mapping": aux_mapping,
        "충원율_추정치_사용": bool(df.attrs.get("충원율_추정치_사용", False)),
        "전형구분_존재": bool(df.attrs.get("전형구분_존재", False)),
    }
    return df, debug_info


# =========================================================================
# 7. Streamlit 앱
# =========================================================================
st.set_page_config(page_title="AI 대학 합격 가능성 분석기", layout="centered")
st.title("🎓 AI 대학 합격 가능성 분석기")
st.caption("제목이 여러 줄이거나 병합 셀·멀티 헤더가 있는 대학 공식 입시결과 엑셀도 자동으로 인식합니다.")

# 세션 상태 초기화 (분석 결과를 화면 갱신 후에도 유지하기 위함)
if "analyzed" not in st.session_state:
    st.session_state.analyzed = False
if "analysis_result" not in st.session_state:
    st.session_state.analysis_result = {}
if "last_file_id" not in st.session_state:
    st.session_state.last_file_id = None


def reset_analysis():
    """학과나 내신 등급이 바뀌면 이전 분석 결과를 숨겨서 혼선을 방지."""
    st.session_state.analyzed = False


uploaded_file = st.file_uploader("전년도 입시결과(Excel)를 업로드하세요", type=["xlsx"])

if uploaded_file is not None:
    # 파일이 바뀌면 이전 분석 결과 초기화
    file_id = getattr(uploaded_file, "file_id", uploaded_file.name)
    if st.session_state.last_file_id != file_id:
        st.session_state.analyzed = False
        st.session_state.last_file_id = file_id

    # ---- 엑셀 자동 파싱 (예외처리 포함) ----
    try:
        df, debug_info = parse_admission_excel(uploaded_file)
    except ValueError as e:
        st.error(f"엑셀 분석 중 오류가 발생했습니다: {e}")
        st.stop()
    except Exception as e:
        st.error(f"예상하지 못한 오류가 발생했습니다: {e}")
        st.stop()

    st.success("입시결과를 자동으로 인식했습니다.")

    # 어떤 행/컬럼을 어떻게 인식했는지 사용자가 확인할 수 있도록 표시
    with st.expander("🔎 엑셀 자동 인식 결과 확인하기"):
        st.write(f"- 데이터 시작 행(엑셀 기준): **{debug_info['data_start_row_in_excel']}행**")
        st.write("- 인식된 원본 컬럼명(멀티헤더 결합 결과):")
        st.write(debug_info["detected_column_names"])
        st.write("- 표준 컬럼 매칭 결과 (원본 컬럼 인덱스, 0부터 시작):")
        st.write(debug_info["column_mapping"])

    # 누락된 선택 항목(경쟁률/충원율)이 있으면 안내만 하고 진행
    optional_missing = [
        c for c in ["경쟁률", "충원율"] if debug_info["column_mapping"].get(c) is None
    ]
    if debug_info.get("충원율_추정치_사용"):
        # 충원율 컬럼이 없어서 "모집인원"과 "충원 최종 예비순위"로 역산한 경우
        optional_missing = [c for c in optional_missing if c != "충원율"]
        st.info("엑셀에 '충원율' 컬럼이 없어, 모집인원 대비 충원 최종 예비순위로 충원율을 추정 계산했습니다.")
    if optional_missing:
        st.warning(f"다음 항목은 엑셀에서 찾지 못해 분석에서 빈 값으로 처리됩니다: {optional_missing}")

    with st.expander("📊 정리된 데이터 확인하기"):
        st.dataframe(df)

    has_admission_type = debug_info.get("전형구분_존재", False)

    # 전형구분(학생부교과/학생부종합/논술 등)이 있으면, 전형을 먼저 고른 뒤
    # 학과를 고르게 해서 같은 학과라도 전형별로 다른 합격선을 정확히 구분한다.
    # 기본값은 "전체 보기"로 두어, 처음 화면에서 특정 전형(학과가 1~2개뿐인 특별전형 등)으로
    # 갑자기 좁혀져서 학과가 거의 안 보이는 문제를 방지한다.
    if has_admission_type:
        ALL_TYPES_LABEL = "전체 보기"
        admission_types = [ALL_TYPES_LABEL] + sorted(
            t for t in df["전형구분"].dropna().unique() if t != ""
        )
        admission_type = st.selectbox(
            "전형구분", admission_types, on_change=reset_analysis, key="admission_type"
        )
        if admission_type == ALL_TYPES_LABEL:
            filtered_df = df
        else:
            filtered_df = df[df["전형구분"] == admission_type]
    else:
        admission_type = None
        filtered_df = df

    departments = sorted(filtered_df["학과"].dropna().unique())
    st.caption(f"현재 선택 기준으로 {len(departments)}개 학과가 검색되었습니다.")

    department = st.selectbox(
        "지원 학과", departments, on_change=reset_analysis, key="department"
    )

    my_grade = st.number_input(
        "내신 등급",
        min_value=1.0,
        max_value=9.0,
        value=2.5,
        step=0.1,
        on_change=reset_analysis,
        key="my_grade",
    )

    # 70%cut을 곧바로 "여기 넘으면 무조건 낮음"으로 딱 자르지 않고,
    # cut보다 조금 더 안 좋은 등급까지는 "보통"으로 봐줄 여유 구간을 사용자가 직접 조절할 수 있게 한다.
    grade_buffer = st.slider(
        "최종합격등급(70%cut) 여유 폭",
        min_value=0.0,
        max_value=1.0,
        value=0.3,
        step=0.1,
        on_change=reset_analysis,
        key="grade_buffer",
        help="70%cut보다 등급이 이 값만큼 더 낮아도(안 좋아도) '보통'으로 판단합니다.",
    )

    if st.button("합격 가능성 분석"):
        matched = filtered_df[filtered_df["학과"] == department]

        if len(matched) > 1:
            st.warning(
                f"'{department}'에 해당하는 데이터가 {len(matched)}건 있습니다. "
                "첫 번째 행을 기준으로 분석합니다."
            )
        row = matched.iloc[0]

        avg = row["평균등급"]
        final = row["최종합격등급"]
        rate = row["경쟁률"]
        add = row["충원율"]

        if pd.isna(avg) or pd.isna(final):
            st.error("선택하신 학과의 등급 데이터(평균등급 또는 최종합격등급)가 유실되어 분석할 수 없습니다.")
            st.session_state.analyzed = False
        else:
            if my_grade <= avg:
                result = "🟢 매우 높음"
                comment = "평균 합격등급보다 우수합니다. 안정적인 지원이 가능합니다."
            elif my_grade <= final + grade_buffer:
                result = "🟡 보통"
                if my_grade <= final:
                    comment = "최종 합격컷(70%cut) 범위 내에 있습니다. 추가합격을 노려볼 수 있습니다."
                else:
                    comment = (
                        "70%cut보다는 살짝 낮지만, 70%cut은 실제 최종 합격선보다 "
                        "다소 우수하게 표시되는 경향이 있어 여유 폭 내에서는 도전해볼 만합니다."
                    )
            else:
                result = "🔴 낮음"
                comment = "70%cut과 여유 폭을 감안해도 낮습니다. 상향 지원에 해당합니다."

            avg_gap = round(my_grade - avg, 2)
            final_gap = round(my_grade - final, 2)
            avg_gap_text = f"평균과 {abs(avg_gap)}등급 " + ("차이(우수)" if avg_gap < 0 else "차이(낮음)" if avg_gap > 0 else "동일")
            final_gap_text = f"최종컷과 {abs(final_gap)}등급 " + ("차이(우수)" if final_gap < 0 else "차이(낮음)" if final_gap > 0 else "동일")

            st.session_state.analysis_result = {
                "department": department,
                "admission_type": admission_type if admission_type != "전체 보기" else None,
                "result": result,
                "comment": comment,
                "avg": avg,
                "final": final,
                "grade_buffer": grade_buffer,
                "my_grade": my_grade,
                "avg_gap_text": avg_gap_text,
                "final_gap_text": final_gap_text,
                "rate": rate,
                "add": add,
            }
            st.session_state.analyzed = True

    if st.session_state.analyzed:
        res = st.session_state.analysis_result

        st.write("---")
        st.header("🔍 분석 결과")
        caption_text = res["department"]
        if res.get("admission_type"):
            caption_text = f"{res['admission_type']} · {caption_text}"
        st.caption(caption_text)
        st.metric("합격 가능성", res["result"])

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("📊 나의 성적 비교")
            st.write(f"내신 등급 : **{res['my_grade']}**")
            st.write(f"평균 합격등급 : **{res['avg']}** ({res['avg_gap_text']})")
            st.write(f"최종 합격등급(70%cut) : **{res['final']}** ({res['final_gap_text']})")
            if res["grade_buffer"] > 0:
                st.caption(f"여유 폭 {res['grade_buffer']}등급 적용 → 실질 기준선 {round(res['final'] + res['grade_buffer'], 2)}")

        with col2:
            st.subheader("📈 학과 입시 지표")
            rate_text = f"{res['rate']} : 1" if pd.notna(res["rate"]) else "정보 없음"
            add_text = f"{res['add']}%" if pd.notna(res["add"]) else "정보 없음"
            st.write(f"경쟁률 : **{rate_text}**")
            st.write(f"충원율(예비 번호) : **{add_text}**")

        st.info(res["comment"])

        if pd.notna(res["add"]) and res["add"] >= 100:
            st.success("💡 이 학과는 전년도 충원율이 100% 이상으로, 한 바퀴 이상 예비번호가 돌았습니다. 추가합격 변수를 긍정적으로 고려해보세요.")

        st.caption("⚠️ 본 분석은 전년도 데이터를 기반으로 한 참고용 정보이며, 실제 당해 년도 수험생 지원 성향 및 모집 인원에 따라 결과가 달라질 수 있습니다.")